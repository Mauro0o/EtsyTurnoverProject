"""
exporter.py - SQLite and Excel export for the Etsy Turnover Scraper.

SQLiteExporter:
  - Creates tables (sold_listings, active_listings, matched_turnover)
  - Upserts records idempotently
  - Reads rows back for export

Schema notes (v2 – preserve_all sold dedup mode):
  sold_listings.sold_row_id  is the PRIMARY KEY (replaces the old composite PK
      on listing_id+domain+shop_name).  This allows multiple rows for the same
      listing_id to coexist as distinct sale events.
      sold_row_id is deterministic: "{listing_id}|{shop_name}|{domain}|p{page}|i{idx}"
      so re-runs replace (not duplicate) existing rows.

  matched_turnover now has a sales_count column and estimated_turnover reflects
      price × sales_count.  One row per unique sold listing_id per run.

  IMPORTANT: If you have an existing database created before this version you
  must DELETE it so the new schema is created from scratch.  SQLite does not
  support changing a table's primary key in-place.

ExcelExporter:
  - Writes a 3-sheet workbook (sold_listings, active_listings, matched_turnover)
  - Styled header row
  - Numeric columns preserved as floats
  - Auto-sized columns (capped at 60 characters)

export_csv (standalone):
  - Writes any list[dict] to a UTF-8 CSV file
"""

from __future__ import annotations

import csv
import logging
import sqlite3
from pathlib import Path
from typing import Optional, Sequence

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models import ActiveListing, MatchedTurnoverRow, SoldListing

logger = logging.getLogger(__name__)

# Columns that should remain numeric in Excel (not coerced to string).
_NUMERIC_COLS: frozenset[str] = frozenset({
    "price",
    "estimated_price",
    "estimated_turnover",
    "sales_count",
    "sold_flag",
    "matched_flag",
    "listing_position_on_page",
    "sold_page_number",
    "storefront_page_number",
})


# ---------------------------------------------------------------------------
# SQLite Exporter
# ---------------------------------------------------------------------------


class SQLiteExporter:
    """Manages all database writes for listings and turnover data."""

    def __init__(self, db_path: Path) -> None:
        self.db_path = db_path
        self._conn: Optional[sqlite3.Connection] = None

    def connect(self) -> None:
        """Open connection and create tables if they do not yet exist."""
        self._conn = sqlite3.connect(str(self.db_path))
        self._conn.execute("PRAGMA journal_mode=WAL")
        self._conn.execute("PRAGMA synchronous=NORMAL")
        self._create_tables()
        logger.debug("SQLite connected: %s", self.db_path)

    def close(self) -> None:
        """Commit any pending writes and close the connection."""
        if self._conn:
            try:
                self._conn.commit()
            except Exception:
                pass
            self._conn.close()
            self._conn = None

    # ------------------------------------------------------------------
    # Schema
    # ------------------------------------------------------------------

    def _create_tables(self) -> None:
        assert self._conn is not None
        self._conn.executescript(
            """
            -- sold_listings: sold_row_id is the PK so repeated listing_ids
            -- (= multiple sales of the same item) coexist as distinct rows.
            CREATE TABLE IF NOT EXISTS sold_listings (
                sold_row_id             TEXT NOT NULL PRIMARY KEY,
                listing_id              TEXT NOT NULL,
                scrape_timestamp        TEXT,
                domain                  TEXT,
                shop_name               TEXT,
                shop_id                 TEXT,
                sold_page_url           TEXT,
                sold_page_number        INTEGER,
                listing_url             TEXT,
                product_title           TEXT,
                image_url               TEXT,
                sold_flag               INTEGER DEFAULT 1,
                card_text_status        TEXT,
                sold_price_raw          TEXT,
                currency                TEXT,
                extraction_notes        TEXT,
                raw_html_snapshot_path  TEXT
            );

            CREATE TABLE IF NOT EXISTS active_listings (
                listing_id               TEXT NOT NULL,
                scrape_timestamp         TEXT,
                domain                   TEXT,
                shop_name                TEXT,
                shop_id                  TEXT,
                storefront_page_url      TEXT,
                storefront_page_number   INTEGER,
                listing_url              TEXT,
                product_title            TEXT,
                image_url                TEXT,
                price                    REAL,
                currency                 TEXT,
                availability             TEXT,
                availability_raw         TEXT,
                listing_position_on_page INTEGER,
                extraction_notes         TEXT,
                raw_html_snapshot_path   TEXT,
                PRIMARY KEY (listing_id, domain, shop_name)
            );

            -- matched_turnover: one row per unique sold listing_id per run.
            -- sales_count = number of times that listing_id appeared in sold rows.
            -- estimated_turnover = estimated_price * sales_count.
            CREATE TABLE IF NOT EXISTS matched_turnover (
                id                  INTEGER PRIMARY KEY AUTOINCREMENT,
                scrape_timestamp    TEXT,
                domain              TEXT,
                shop_name           TEXT,
                sold_listing_id     TEXT NOT NULL,
                active_listing_id   TEXT,
                match_type          TEXT,
                sold_title          TEXT,
                active_title        TEXT,
                estimated_price     REAL,
                currency            TEXT,
                sales_count         INTEGER DEFAULT 1,
                estimated_turnover  REAL,
                matched_flag        INTEGER DEFAULT 0,
                notes               TEXT,
                UNIQUE (sold_listing_id, domain, shop_name)
            );
            """
        )
        self._conn.commit()

    # ------------------------------------------------------------------
    # Upsert methods
    # ------------------------------------------------------------------

    def upsert_sold_listings(self, listings: Sequence[SoldListing]) -> int:
        """
        Insert or replace sold listings, keyed on sold_row_id.

        Because sold_row_id is deterministic, re-running the scraper replaces
        existing rows rather than duplicating them.

        Returns the number of rows processed.
        """
        if not listings:
            return 0
        assert self._conn is not None
        rows = [_sold_to_row(s) for s in listings]
        self._conn.executemany(
            """
            INSERT OR REPLACE INTO sold_listings VALUES
            (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            rows,
        )
        self._conn.commit()
        logger.debug("Upserted %d sold listings", len(rows))
        return len(rows)

    def upsert_active_listings(self, listings: Sequence[ActiveListing]) -> int:
        """
        Insert or replace active listings.  Idempotent on (listing_id, domain, shop_name).
        """
        if not listings:
            return 0
        assert self._conn is not None
        rows = [_active_to_row(a) for a in listings]
        self._conn.executemany(
            """
            INSERT OR REPLACE INTO active_listings VALUES
            (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            rows,
        )
        self._conn.commit()
        logger.debug("Upserted %d active listings", len(rows))
        return len(rows)

    def upsert_matched_turnover(self, rows: Sequence[MatchedTurnoverRow]) -> int:
        """
        Insert or replace matched turnover rows.
        Idempotent on (sold_listing_id, domain, shop_name).
        """
        if not rows:
            return 0
        assert self._conn is not None
        data = [_matched_to_row(m) for m in rows]
        self._conn.executemany(
            """
            INSERT OR REPLACE INTO matched_turnover
            (scrape_timestamp, domain, shop_name, sold_listing_id, active_listing_id,
             match_type, sold_title, active_title, estimated_price, currency,
             sales_count, estimated_turnover, matched_flag, notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """,
            data,
        )
        self._conn.commit()
        logger.debug("Upserted %d matched turnover rows", len(data))
        return len(data)

    # ------------------------------------------------------------------
    # Read-back (used by exporters)
    # ------------------------------------------------------------------

    def fetch_all_sold(self) -> list[dict]:
        assert self._conn is not None
        cur = self._conn.execute("SELECT * FROM sold_listings")
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]

    def fetch_all_active(self) -> list[dict]:
        assert self._conn is not None
        cur = self._conn.execute("SELECT * FROM active_listings")
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]

    def fetch_all_matched(self) -> list[dict]:
        assert self._conn is not None
        cur = self._conn.execute("SELECT * FROM matched_turnover")
        cols = [d[0] for d in cur.description]
        return [dict(zip(cols, row)) for row in cur.fetchall()]


# ---------------------------------------------------------------------------
# Excel Exporter
# ---------------------------------------------------------------------------


class ExcelExporter:
    """Exports listing data to a 3-sheet Excel workbook using openpyxl."""

    _HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    _HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    _ALT_ROW_FILL = PatternFill("solid", fgColor="E8F0FE")  # light blue for even rows

    def export(
        self,
        sold_rows: list[dict],
        active_rows: list[dict],
        matched_rows: list[dict],
        output_path: Path,
    ) -> None:
        """
        Write a workbook with three sheets to *output_path*.

        Overwrites any existing file at that path.
        """
        wb = openpyxl.Workbook()
        # Remove the default empty sheet created by openpyxl.
        default_sheet = wb.active
        if default_sheet:
            wb.remove(default_sheet)

        self._write_sheet(wb, "sold_listings", sold_rows)
        self._write_sheet(wb, "active_listings", active_rows)
        self._write_sheet(wb, "matched_turnover", matched_rows)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        logger.info("Excel workbook saved: %s", output_path)

    def _write_sheet(
        self,
        wb: openpyxl.Workbook,
        sheet_name: str,
        rows: list[dict],
    ) -> None:
        ws = wb.create_sheet(title=sheet_name)

        if not rows:
            ws.append(["(no data)"])
            return

        headers = list(rows[0].keys())
        ws.append(headers)

        # Style header row.
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = self._HEADER_FILL
            cell.font = self._HEADER_FONT

        # Write data rows with type-aware cell values.
        for row_idx, row_dict in enumerate(rows, start=2):
            row_values = []
            for col_name, val in row_dict.items():
                if col_name in _NUMERIC_COLS and val is not None:
                    try:
                        row_values.append(float(val))
                    except (ValueError, TypeError):
                        row_values.append(val)
                else:
                    # Ensure None is written as empty string in Excel.
                    row_values.append(val if val is not None else "")
            ws.append(row_values)

        # Auto-size columns based on header + data content (capped at 60).
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header))
            for row_dict in rows:
                cell_val = row_dict.get(header, "") or ""
                max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        # Freeze the header row.
        ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------


def export_csv(rows: list[dict], path: Path) -> None:
    """
    Write *rows* to a UTF-8 CSV file at *path*.

    Creates parent directories if needed.  Silently skips if rows is empty.
    """
    if not rows:
        logger.debug("No rows to write to CSV: %s", path)
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)
    logger.info("CSV exported: %s (%d rows)", path, len(rows))


# ---------------------------------------------------------------------------
# Row serialisers
# ---------------------------------------------------------------------------


def _sold_to_row(s: SoldListing) -> tuple:
    return (
        s.sold_row_id,
        s.listing_id,
        s.scrape_timestamp,
        s.domain,
        s.shop_name,
        s.shop_id,
        s.sold_page_url,
        s.sold_page_number,
        s.listing_url,
        s.product_title,
        s.image_url,
        int(s.sold_flag),
        s.card_text_status,
        s.sold_price_raw,
        s.currency,
        s.extraction_notes,
        s.raw_html_snapshot_path,
    )


def _active_to_row(a: ActiveListing) -> tuple:
    return (
        a.listing_id,
        a.scrape_timestamp,
        a.domain,
        a.shop_name,
        a.shop_id,
        a.storefront_page_url,
        a.storefront_page_number,
        a.listing_url,
        a.product_title,
        a.image_url,
        a.price,
        a.currency,
        a.availability,
        a.availability_raw,
        a.listing_position_on_page,
        a.extraction_notes,
        a.raw_html_snapshot_path,
    )


def _matched_to_row(m: MatchedTurnoverRow) -> tuple:
    return (
        m.scrape_timestamp,
        m.domain,
        m.shop_name,
        m.sold_listing_id,
        m.active_listing_id,
        m.match_type,
        m.sold_title,
        m.active_title,
        m.estimated_price,
        m.currency,
        m.sales_count,
        m.estimated_turnover,
        m.matched_flag,
        m.notes,
    )
